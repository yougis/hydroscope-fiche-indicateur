#!/usr/bin/env python3
"""
Génération Typst des fiches indicateurs HydroScope.
"""
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from pathlib import Path


XLSX = Path("fiches indicateurs.xlsx")
OUTPUT_XLSX = "liste_flat_indicateurs.xlsx"

# ── Palette fixe par famille_indicateur (tint 0.0, couleur pleine) ──────────
FAMILLE_BASE_COLORS = {
    'Enjeu':         '2b5e8c',   # Bleu
    'Menace':      'faa51a',   # Orange
    'Vulnérabilité': '318d44',   # Vert
    'Contexte':      'FFC000',   # Jaune/Or
}

# ── Palette fixe par Nom_theme ───────────────────────────────────────────────
THEME_COLORS = {
    'Enjeux AEP':                  '4472C4',   # Bleu marine profond
    'Enjeux Environnementaux':     '3bc29f',   # Bleu ciel moyen
    'MENACES NATURELLES': 'ED7D31',   # Brun-orange foncé
    'MENACES ANTHROPIQUES':      'C55A11',   # Orange foncé
    'Menaces Quantitatives':     'ED7D31',   # Orange vif
    'Menaces Qualitatives':      'F4B183',   # Orange pâle
    'Vulnérabilité Intrinsèque':   '41af2f',   # Vert forêt
}

# ── Plage de tint pour la variation des groupes au sein d'un thème ──────────
# Les groupes varient dans [tint_min, tint_max] autour de la couleur du thème
GROUPE_TINT_RANGE = (0.2, 0.6)

DEFAULT_HEX = '7F7F7F'


def make_fill(hex_color: str, tint: float = 0.0) -> PatternFill:
    """Crée un PatternFill openpyxl avec teinte bornée dans [0, 0.99]."""
    c = Color(rgb=hex_color)
    c.tint = max(0.0, min(tint, 0.99))
    return PatternFill(start_color=c, end_color=c, fill_type='solid')


def interpolate(rank: int, total: int, v_min: float, v_max: float) -> float:
    """Interpolation linéaire dans [v_min, v_max] selon le rang (0-based)."""
    if total <= 1:
        return v_min
    return v_min + (v_max - v_min) * (rank / (total - 1))


def build_fills_cache(df: pd.DataFrame) -> dict:
    """
    Pré-calcule tous les PatternFill selon la hiérarchie :
      famille  → couleur fixe FAMILLE_BASE_COLORS, tint 0.0
      theme    → couleur fixe THEME_COLORS, tint 0.0
      groupe   → couleur du thème parent + tint variant dans GROUPE_TINT_RANGE
                 rang calculé au sein du thème uniquement

    Retourne :
    {
        famille_str: {
            'fill': PatternFill,
            'themes': {
                theme_str: {
                    'fill':    PatternFill,
                    'groupes': { groupe_str: PatternFill }
                }
            }
        }
    }
    """
    cache = {}
    gr_min, gr_max = GROUPE_TINT_RANGE

    for famille in df['famille_indicateur'].unique():
        famille_hex  = FAMILLE_BASE_COLORS.get(str(famille), DEFAULT_HEX)
        famille_fill = make_fill(famille_hex, 0.0)
        df_fam       = df[df['famille_indicateur'] == famille]

        themes_dict = {}
        for theme in df_fam['Nom_theme'].unique():
            theme_hex  = THEME_COLORS.get(str(theme), DEFAULT_HEX)
            theme_fill = make_fill(theme_hex, 0.0)   # couleur pleine fixe

            # Groupes : variation de tint dans GROUPE_TINT_RANGE
            # sur la couleur du thème parent, rang local au thème
            groupes = list(
                df_fam[df_fam['Nom_theme'] == theme]['groupe'].unique()
            )
            total_groupes = len(groupes)
            groupes_dict  = {}

            for gr_rank, groupe in enumerate(groupes):
                tint_groupe = interpolate(gr_rank, total_groupes, gr_min, gr_max)
                groupes_dict[str(groupe)] = make_fill(theme_hex, tint_groupe)

            themes_dict[str(theme)] = {
                'fill':    theme_fill,
                'groupes': groupes_dict,
            }

        cache[str(famille)] = {
            'fill':   famille_fill,
            'themes': themes_dict,
        }

    return cache


def main():
    print("🚀 Génération d'une liste des indicateurs HydroScope...")

    try:
        tabs = pd.read_excel(XLSX, sheet_name=None)
    except Exception as exc:
        print(f'❌ Impossible de charger {XLSX}: {exc}')
        return

    df_ind       = tabs['indicateurs']
    df_group     = tabs['groupes']
    df_rel_group = tabs['relation groupe objectifs']
    df_theme     = tabs['themes']
    df_famille   = tabs['Familles']

    for df in [df_ind, df_group, df_rel_group, df_theme, df_famille]:
        df.columns = df.columns.str.strip()

    # 1. Filtrage sur les indicateurs actifs
    df_ind = df_ind[df_ind['actif'].astype(str).str.lower() == 'oui']

    # 2. Jointures relationnelles
    df_flat = pd.merge(df_ind, df_rel_group, on='id_indicateur', how='left', suffixes=('', '_rel'))
    df_flat = pd.merge(df_flat, df_group,     on='id_groupe',     how='left', suffixes=('', '_group'))
    if 'theme_id' in df_theme.columns:
        df_flat = pd.merge(df_flat, df_theme, left_on='Nom_theme', right_on='Nom_theme', how='left', suffixes=('', '_theme'))

    # 3. Sélection et ordre des colonnes
    colonnes_a_garder = [
        'id_indicateur',
        'famille_indicateur',
        'Nom_theme',
        'groupe',
        'nom_indicateur',
        'description_indicateur',
        'objectif',
        'priorite'
    ]
    colonnes_finales = [col for col in colonnes_a_garder if col in df_flat.columns]
    df_flat = df_flat[colonnes_finales]
    df_flat = df_flat.sort_values(
        by=['famille_indicateur', 'Nom_theme', 'groupe']
    ).reset_index(drop=True)

    # 4. Pré-calcul du cache de fills
    fills_cache = build_fills_cache(df_flat)

    # 5. Index Excel des colonnes (1-based)
    def col_idx(name):
        return colonnes_finales.index(name) + 1 if name in colonnes_finales else None

    idx_famille = col_idx('famille_indicateur')
    idx_theme   = col_idx('Nom_theme')
    idx_groupe  = col_idx('groupe')

    # Colonnes non-hiérarchiques (tout sauf id_indicateur et les 3 niveaux)
    cols_exclues = {'famille_indicateur', 'id_indicateur', 'Nom_theme', 'groupe'}
    idx_autres = [
        colonnes_finales.index(c) + 1
        for c in colonnes_finales
        if c not in cols_exclues
    ]

    # 6. Construction du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicateurs Actifs"
    ws.views.sheetView[0].showGridLines = True

    for r in dataframe_to_rows(df_flat, index=False, header=True):
        ws.append(r)

    # 7. En-têtes
    header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    header_font = openpyxl.styles.Font(color='FFFFFF', bold=True)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font

    # 8. Coloration des lignes de données
    for row_idx in range(2, ws.max_row + 1):
        famille_val = str(ws.cell(row=row_idx, column=idx_famille).value) if idx_famille else None
        theme_val   = str(ws.cell(row=row_idx, column=idx_theme).value)   if idx_theme   else None
        groupe_val  = str(ws.cell(row=row_idx, column=idx_groupe).value)  if idx_groupe  else None

        famille_entry = fills_cache.get(famille_val)
        if not famille_entry:
            continue

        theme_entry = famille_entry['themes'].get(theme_val)
        groupe_fill = theme_entry['groupes'].get(groupe_val) if theme_entry else None

        # Colonne famille → couleur fixe pleine
        if idx_famille:
            ws.cell(row=row_idx, column=idx_famille).fill = famille_entry['fill']

        # Colonne Nom_theme → couleur fixe pleine du thème
        if idx_theme and theme_entry:
            ws.cell(row=row_idx, column=idx_theme).fill = theme_entry['fill']

        # Colonne groupe + toutes les autres → couleur thème + tint progressif
        if groupe_fill:
            if idx_groupe:
                ws.cell(row=row_idx, column=idx_groupe).fill = groupe_fill
            for c_idx in idx_autres:
                ws.cell(row=row_idx, column=c_idx).fill = groupe_fill

    # 9. Ajustement automatique des largeurs
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(OUTPUT_XLSX)
    print(f'✨ Fichier généré avec succès : {OUTPUT_XLSX}')


if __name__ == '__main__':
    main()